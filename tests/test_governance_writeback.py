"""
Tests for utils.governance._append_to_parameter_sheet — the only governance
write-back path (the read/delta side is covered by test_delta_detection).

Builds a synthetic master_monitoring.xlsx (PARAMETER sheet: col A = PM,
col D = Anchor) in a tmp dir, appends net-new entities, and verifies the
correct UNASSIGNED / UNMAPPED_ANCHOR defaults, idempotency, and error paths.

Run:
    pytest tests/test_governance_writeback.py -v
"""

import os
import sys

import openpyxl
import pytest

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from utils.governance import _append_to_parameter_sheet, _read_master_parameter


def _make_param_book(path, data_rows, *, sheet="PARAMETER"):
    """Create an .xlsx with a header row + (pm, anchor) data rows in col A / col D."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(row=1, column=1).value = "PM"
    ws.cell(row=1, column=2).value = "NO"
    ws.cell(row=1, column=3).value = "KODE"
    ws.cell(row=1, column=4).value = "ANCHOR"
    for i, (pm, anchor) in enumerate(data_rows, start=2):
        ws.cell(row=i, column=1).value = pm
        ws.cell(row=i, column=4).value = anchor
    wb.save(path)
    return path


def _pairs(path):
    """Return {(anchor, pm)} as read back from the PARAMETER sheet."""
    df = _read_master_parameter(path)
    return {(r["Anchor"], r["PM"]) for _, r in df.iterrows()}


class TestAppendToParameterSheet:
    def test_appends_new_anchor_and_pm_with_defaults(self, tmp_path):
        path = _make_param_book(tmp_path / "master_monitoring.xlsx", [("ALICE", "ACME")])

        _append_to_parameter_sheet(str(path), ["NEWANCHOR"], ["BOB"])

        pairs = _pairs(path)
        assert ("ACME", "ALICE") in pairs                 # original preserved
        assert ("NEWANCHOR", "UNASSIGNED") in pairs       # new anchor -> PM UNASSIGNED
        assert ("UNMAPPED_ANCHOR", "BOB") in pairs        # new pm -> Anchor UNMAPPED_ANCHOR

    def test_is_idempotent_on_rerun(self, tmp_path):
        path = _make_param_book(tmp_path / "master_monitoring.xlsx", [("ALICE", "ACME")])

        _append_to_parameter_sheet(str(path), ["NEWANCHOR"], ["BOB"])
        after_first = len(_read_master_parameter(path))
        _append_to_parameter_sheet(str(path), ["NEWANCHOR"], ["BOB"])
        after_second = len(_read_master_parameter(path))

        assert after_first == after_second  # no duplicate rows

    def test_missing_sheet_raises_value_error(self, tmp_path):
        path = tmp_path / "no_param.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "SOMETHING_ELSE"
        wb.save(path)
        with pytest.raises(ValueError, match="PARAMETER"):
            _append_to_parameter_sheet(str(path), ["X"], [])

    def test_missing_file_raises_file_not_found(self, tmp_path):
        missing = tmp_path / "does_not_exist.xlsx"
        with pytest.raises(FileNotFoundError):
            _append_to_parameter_sheet(str(missing), ["X"], [])
