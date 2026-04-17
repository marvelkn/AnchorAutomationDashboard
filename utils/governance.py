"""
utils/governance.py
-------------------
Pure-Python governance gate for the ETL pipeline.

Detects new (unknown) Anchors and PMs by comparing the TARGET table
in an uploaded SQLite database against the PARAMETER sheet in the
master monitoring Excel file.  No Streamlit dependency — safe to
import in tests, CLI scripts, and background workers.

Public API
----------
_norm_text(value)                             -> str | None
_read_target_entities(db_path)                -> pd.DataFrame
_read_master_parameter(path_mon)              -> pd.DataFrame
_detect_governance_delta(db_path, path_mon)   -> dict
_compute_db_signature(db_path)                -> str
_append_to_parameter_sheet(path_mon, approved_anchors, approved_pms)
_write_governance_audit(audit_path, decisions)
"""

from __future__ import annotations

import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl


# ── Text normalisation ────────────────────────────────────────────────────────

def _norm_text(value) -> Optional[str]:
    """Strip whitespace; return None for NaN / empty / null-like strings."""
    if pd.isna(value):
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned.lower() in {"nan", "none", "null"}:
        return None
    return cleaned


# ── Readers ───────────────────────────────────────────────────────────────────

def _read_target_entities(db_path: str) -> pd.DataFrame:
    """
    Read (MERCHANT_GROUP, PM) pairs from the TARGET table in *db_path*.

    Returns an empty DataFrame with columns ['Anchor', 'PM'] on any error.
    """
    if not os.path.exists(db_path):
        return pd.DataFrame(columns=["Anchor", "PM"])
    conn = sqlite3.connect(db_path)
    try:
        query = """
            SELECT
                TRIM(MERCHANT_GROUP) AS Anchor,
                TRIM(PM)             AS PM
            FROM TARGET
            WHERE MERCHANT_GROUP IS NOT NULL OR PM IS NOT NULL
        """
        return pd.read_sql_query(query, conn)
    except Exception:
        return pd.DataFrame(columns=["Anchor", "PM"])
    finally:
        conn.close()


def _read_master_parameter(path_mon: str) -> pd.DataFrame:
    """
    Read known (Anchor, PM) pairs from the PARAMETER sheet in
    *path_mon* (master_monitoring.xlsx).

    Column A = PM, Column D = Anchor (rows 2 onwards).
    Returns an empty DataFrame with columns ['Anchor', 'PM'] on any error.
    """
    if not os.path.exists(path_mon):
        return pd.DataFrame(columns=["Anchor", "PM"])
    try:
        wb = openpyxl.load_workbook(path_mon, data_only=True)
        if "PARAMETER" not in wb.sheetnames:
            return pd.DataFrame(columns=["Anchor", "PM"])
        ws = wb["PARAMETER"]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            pm_val     = _norm_text(ws.cell(row=row_idx, column=1).value)  # col A
            anchor_val = _norm_text(ws.cell(row=row_idx, column=4).value)  # col D
            if pm_val or anchor_val:
                rows.append({"Anchor": anchor_val, "PM": pm_val})
        return pd.DataFrame(rows, columns=["Anchor", "PM"])
    except Exception:
        return pd.DataFrame(columns=["Anchor", "PM"])


# ── Core delta detection ──────────────────────────────────────────────────────

def _detect_governance_delta(db_path: str, path_mon: str) -> dict:
    """
    Compare entities in the uploaded DB against the master PARAMETER sheet.

    Returns
    -------
    {
        "new_anchors":        list[str],   # Anchors in DB but not in master
        "new_pms":            list[str],   # PMs in DB but not in master
        "impact_anchor_rows": int,         # Rows in DB that contain a new Anchor
        "impact_pm_rows":     int,         # Rows in DB that contain a new PM
    }
    This function is READ-ONLY — it never modifies any file or database.
    """
    uploaded_df = _read_target_entities(db_path)
    master_df   = _read_master_parameter(path_mon)

    up_anchor_series = (
        uploaded_df["Anchor"].map(_norm_text).dropna()
        if "Anchor" in uploaded_df.columns
        else pd.Series(dtype="object")
    )
    up_pm_series = (
        uploaded_df["PM"].map(_norm_text).dropna()
        if "PM" in uploaded_df.columns
        else pd.Series(dtype="object")
    )
    m_anchor_series = (
        master_df["Anchor"].map(_norm_text).dropna()
        if "Anchor" in master_df.columns
        else pd.Series(dtype="object")
    )
    m_pm_series = (
        master_df["PM"].map(_norm_text).dropna()
        if "PM" in master_df.columns
        else pd.Series(dtype="object")
    )

    uploaded_anchors = {x for x in up_anchor_series.tolist() if x}
    uploaded_pms     = {x for x in up_pm_series.tolist()     if x}
    master_anchors   = {x for x in m_anchor_series.tolist()  if x}
    master_pms       = {x for x in m_pm_series.tolist()      if x}

    new_anchors = sorted(uploaded_anchors - master_anchors)
    new_pms     = sorted(uploaded_pms     - master_pms)

    impact_anchor_rows = int(up_anchor_series.isin(new_anchors).sum()) if new_anchors else 0
    impact_pm_rows     = int(up_pm_series.isin(new_pms).sum())         if new_pms     else 0

    return {
        "new_anchors":        new_anchors,
        "new_pms":            new_pms,
        "impact_anchor_rows": impact_anchor_rows,
        "impact_pm_rows":     impact_pm_rows,
    }


# ── DB fingerprint ────────────────────────────────────────────────────────────

def _compute_db_signature(db_path: str) -> str:
    """
    Return a lightweight fingerprint of *db_path* based on mtime + size.
    Used to detect whether the DB has changed between gate evaluations.
    """
    if not os.path.exists(db_path):
        return "missing-db"
    stat = os.stat(db_path)
    return f"{int(stat.st_mtime)}-{int(stat.st_size)}"


# ── Write-back helpers (called only after user approval) ──────────────────────

def _append_to_parameter_sheet(
    path_mon: str,
    approved_anchors: list,
    approved_pms: list,
) -> None:
    """
    Append net-new Anchors and PMs to the PARAMETER sheet in *path_mon*.

    Anchors get PM = "UNASSIGNED"; PMs get Anchor = "UNMAPPED_ANCHOR".
    Auto-numbering formulas (columns B and C) are written for each new row.
    Raises FileNotFoundError / ValueError if the file or sheet is missing.
    """
    if not os.path.exists(path_mon):
        raise FileNotFoundError(f"Master monitoring file not found: {path_mon}")

    wb = openpyxl.load_workbook(path_mon)
    if "PARAMETER" not in wb.sheetnames:
        raise ValueError("Sheet 'PARAMETER' not found in master_monitoring.xlsx")
    ws = wb["PARAMETER"]

    existing_anchor: set = set()
    existing_pm: set     = set()
    max_data_row: int    = 1

    for row_idx in range(2, ws.max_row + 1):
        pm_val     = _norm_text(ws.cell(row=row_idx, column=1).value)
        anchor_val = _norm_text(ws.cell(row=row_idx, column=4).value)
        if pm_val:
            existing_pm.add(pm_val)
            max_data_row = row_idx
        if anchor_val:
            existing_anchor.add(anchor_val)
            max_data_row = row_idx

    for anchor in approved_anchors:
        if anchor in existing_anchor:
            continue
        max_data_row += 1
        ws.cell(row=max_data_row, column=1).value = "UNASSIGNED"
        ws.cell(row=max_data_row, column=4).value = anchor
        ws.cell(row=max_data_row, column=2).value = (
            f"=IF(A{max_data_row}=A{max_data_row-1},B{max_data_row-1}+1,1)"
        )
        ws.cell(row=max_data_row, column=3).value = (
            f"=CONCATENATE(A{max_data_row},B{max_data_row})"
        )
        existing_anchor.add(anchor)

    for pm in approved_pms:
        if pm in existing_pm:
            continue
        max_data_row += 1
        ws.cell(row=max_data_row, column=1).value = pm
        ws.cell(row=max_data_row, column=4).value = "UNMAPPED_ANCHOR"
        ws.cell(row=max_data_row, column=2).value = (
            f"=IF(A{max_data_row}=A{max_data_row-1},B{max_data_row-1}+1,1)"
        )
        ws.cell(row=max_data_row, column=3).value = (
            f"=CONCATENATE(A{max_data_row},B{max_data_row})"
        )
        existing_pm.add(pm)

    wb.save(path_mon)


def _write_governance_audit(audit_path: str, decisions: dict) -> None:
    """
    Append governance decisions to *audit_path* (CSV).

    *decisions* must have keys:
        approved_anchors, ignored_anchors, approved_pms, ignored_pms
    Each value is a list of entity name strings.
    Creates the file (and parent dirs) if it doesn't exist.
    Does nothing if all lists are empty.
    """
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for entity in decisions.get("approved_anchors", []):
        rows.append({"timestamp": now_str, "entity_type": "Anchor", "entity_value": entity, "decision": "approve"})
    for entity in decisions.get("ignored_anchors", []):
        rows.append({"timestamp": now_str, "entity_type": "Anchor", "entity_value": entity, "decision": "ignore"})
    for entity in decisions.get("approved_pms", []):
        rows.append({"timestamp": now_str, "entity_type": "PM",     "entity_value": entity, "decision": "approve"})
    for entity in decisions.get("ignored_pms", []):
        rows.append({"timestamp": now_str, "entity_type": "PM",     "entity_value": entity, "decision": "ignore"})

    if not rows:
        return

    audit_df = pd.DataFrame(rows)
    if os.path.exists(audit_path):
        old        = pd.read_csv(audit_path)
        audit_df   = pd.concat([old, audit_df], ignore_index=True)
    Path(audit_path).parent.mkdir(parents=True, exist_ok=True)
    audit_df.to_csv(audit_path, index=False)
