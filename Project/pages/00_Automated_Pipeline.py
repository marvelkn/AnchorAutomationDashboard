import streamlit as st
import os
import sys
import shutil
from datetime import datetime
import pandas as pd
import sqlite3
from pathlib import Path

import openpyxl

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

from utils.theme import (
    apply_theme, page_header, section_label, pipeline_stepper,
)
from utils.db_connector import fetch_data_from_db, get_db_date_bounds
from utils.backup_manager import rotate_backups, get_available_backups, restore_backup
from utils.pipeline_bg import get_pipeline_status, start_pipeline_background, reset_pipeline_status, is_pipeline_supported
from utils.cloud_db import build_engine, test_connection, read_uploaded_dataframe, upsert_dataframe
from utils.sqlite_to_neon import ingest_sqlite_bytes_to_neon, fetch_recent_ingestion_runs
from utils.master_files_db import sync_all_masters_to_disk

cloud_mode_enabled = bool(os.getenv("DATABASE_URL"))

st.set_page_config(page_title="Automated Pipeline — BTN Anchor", page_icon="🚀", layout="wide")
apply_theme()

BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_UPLOAD_DIR = os.path.join(BASE_DIR, "database")
MASTER_DIR    = os.path.join(BASE_DIR, "data", "master")

PATH_LOCAL_DB = os.path.join(DB_UPLOAD_DIR, "staging.db")
PATH_MID      = os.path.join(MASTER_DIR, "master_mid.xlsx")
PATH_CARD     = os.path.join(MASTER_DIR, "master_card_share.xlsx")
PATH_MON      = os.path.join(MASTER_DIR, "master_monitoring.xlsx")
PATH_BACKUP_DIR = os.path.join(DB_UPLOAD_DIR, "backup")
PATH_GOV_AUDIT = os.path.join(MASTER_DIR, "governance_audit_log.csv")

# Create backup dir if missing
if not os.path.exists(PATH_BACKUP_DIR):
    os.makedirs(PATH_BACKUP_DIR)

# ── Cloud: sync master files from Neon → local disk ───────────────────────────
# This ensures PATH_MON / PATH_CARD / PATH_MID exist for the governance check
# and the local pipeline, even on ephemeral cloud filesystems.
if cloud_mode_enabled:
    try:
        _sync_engine = build_engine()
        sync_all_masters_to_disk(_sync_engine, PATH_MID, PATH_CARD, PATH_MON)
    except Exception:
        pass  # Silent — governance check will use whatever is already on disk


# ── Cloud (Neon): single clean page — no local SQLite / Excel pipeline UI ─────
if cloud_mode_enabled:
    page_header(
        "🚀",
        "Automated pipeline — cloud",
        "Load staging data into Neon (PostgreSQL)",
    )
    st.markdown(
        """<div class="tab-desc">
        <p><b>Choose how you want to load data</b> — both options use the same Neon database:</p>
        <ul style="margin:0.35rem 0 0 1rem;line-height:1.55;">
        <li><b>Full database</b> — Upload your SQLite <code>.db</code> file (e.g. <code>staging.db</code>).
        All business tables are copied into Neon in one step (each table is replaced on load). Use this when you have a full export.</li>
        <li><b>Optional manual file</b> — Upload a single <b>CSV</b> or <b>Excel</b> file to upsert <i>one</i> table only.
        Use this for small corrections without re-uploading the whole database.</li>
        </ul>
        <p style="margin-top:0.65rem;opacity:0.88;font-size:0.92em;">
        <b>Bahasa:</b> <i>(1) Upload <b>.db</b> lengkap = semua tabel masuk Neon sekaligus.
        (2) <b>Opsional</b> — CSV/Excel untuk perbarui <b>satu tabel</b> saja.</i>
        </p>
        <p style="margin-top:0.5rem;opacity:0.85;font-size:0.88em;">
        Phase-1 cloud ingestion only — it does not run the legacy Windows + Excel COM analytics pipeline.</p>
        </div>""",
        unsafe_allow_html=True,
    )

    @st.cache_resource
    def _get_cloud_engine():
        return build_engine()

    try:
        engine = _get_cloud_engine()
        test_connection(engine)
        st.success("Connected to Neon (`DATABASE_URL`).")
    except Exception as conn_err:
        st.error(f"Neon connection failed: {conn_err}")
        engine = None

    # ── 2-TAB LAYOUT: Ingest Data | Audit Log ────────────────────────────────
    cloud_tab_ingest, cloud_tab_audit = st.tabs(["📥  Ingest Data", "📋  Audit Log"])

    with cloud_tab_ingest:
        section_label("A — Full SQLite database")
        neon_schema_ingest = st.text_input(
            "Schema for imported tables",
            value="public",
            key="neon_schema_ingest",
            help="Table names are lowercased in Neon (e.g. ALL_MID → all_mid).",
        )
        cloud_db_upload = st.file_uploader(
            "SQLite file (.db / .sqlite)",
            type=["db", "sqlite"],
            key="cloud_db_full_upload",
            help="Written to a temp file only while reading; data is loaded into Neon.",
        )
        if st.button(
            "Ingest full database to Neon",
            type="primary",
            use_container_width=True,
            disabled=(engine is None),
            key="btn_ingest_full_db",
        ):
            if not cloud_db_upload:
                st.warning("Please upload a .db file first.")
            else:
                # Reserve space BEFORE the action starts — prevents layout jumps
                progress_placeholder = st.empty()
                status_placeholder   = st.empty()
                results_placeholder  = st.empty()

                try:
                    def _on_progress(cur: int, total: int, tbl: str, msg: str):
                        frac = min(cur / max(total, 1), 1.0)
                        with progress_placeholder.container():
                            st.progress(frac, text=(msg[:120] if msg else "…"))
                        with status_placeholder.container():
                            st.info(f"Ingesting table {cur}/{total}: `{tbl}`")

                    with st.spinner("Ingesting SQLite → Neon…"):
                        result = ingest_sqlite_bytes_to_neon(
                            engine,
                            cloud_db_upload.getvalue(),
                            schema=(neon_schema_ingest or "public").strip() or "public",
                            source_filename=getattr(cloud_db_upload, "name", "uploaded.db") or "uploaded.db",
                            progress_callback=_on_progress,
                        )

                    # Clear progress, replace with metric summary
                    progress_placeholder.empty()
                    status_placeholder.empty()

                    _ok     = result.get("tables_ok", 0)
                    _total  = (_ok + result.get("tables_failed", 0))
                    _rows   = result.get("total_rows", 0)
                    _elapsed = result.get("elapsed_seconds", 0)
                    _failed  = result.get("tables_failed", 0)

                    with results_placeholder.container():
                        if result.get("status") == "complete":
                            st.success(f"Ingest complete · run `{result.get('run_id')}`")
                        elif result.get("status") == "partial_error":
                            st.warning(f"Partial success · run `{result.get('run_id')}`")
                        else:
                            st.error(f"Ingest failed: {result.get('error_message', 'unknown')}")

                        m1, m2, m3, m4 = st.columns(4)
                        m1.metric("Tables Ingested", f"{_ok}/{_total}")
                        m2.metric("Total Rows",      f"{_rows:,}")
                        m3.metric("Elapsed Time",    f"{_elapsed:.1f}s")
                        m4.metric("Failed Tables",   _failed,
                                  delta=f"-{_failed}" if _failed else None,
                                  delta_color="inverse")

                        tr = result.get("table_results") or []
                        if tr:
                            st.dataframe(pd.DataFrame(tr), use_container_width=True, hide_index=True)
                        with st.expander("Technical details (JSON)"):
                            st.json(result.get("details") or {})

                except Exception as ingest_err:
                    progress_placeholder.empty()
                    status_placeholder.empty()
                    with results_placeholder.container():
                        st.error(f"Full database ingest failed: {ingest_err}")

        with st.expander("B — Optional: single-table upsert (CSV / Excel)", expanded=False):
        st.caption(
            "Target table must already exist in Neon with a PRIMARY KEY or UNIQUE constraint on your conflict column(s)."
        )
        cloud_upload = st.file_uploader(
            "File",
            type=["csv", "xlsx", "xls"],
            key="cloud_upload",
        )
        u1, u2, u3 = st.columns(3)
        cloud_table = u1.text_input("Target table", value="target")
        cloud_schema = u2.text_input("Schema", value="public")
        cloud_keys_raw = u3.text_input("Conflict key(s)", value="merchant_group,pm")
        if st.button(
            "Run upsert",
            type="primary",
            use_container_width=True,
            disabled=(engine is None),
            key="btn_cloud_upsert",
        ):
            if not cloud_upload:
                st.warning("Please upload a CSV or Excel file first.")
            else:
                prog = st.progress(0.0, text="Starting…")
                try:
                    with st.spinner("Reading file…"):
                        cloud_df = read_uploaded_dataframe(cloud_upload)
                    prog.progress(0.35, text=f"Parsed {len(cloud_df):,} rows")
                    conflict_cols = [x.strip() for x in cloud_keys_raw.split(",") if x.strip()]
                    with st.spinner("Upserting…"):
                        affected = upsert_dataframe(
                            engine=engine,
                            dataframe=cloud_df,
                            table_name=cloud_table.strip(),
                            conflict_columns=conflict_cols,
                            schema=(cloud_schema.strip() or "public"),
                        )
                    prog.progress(1.0, text="Done")
                    st.success(
                        f"Upsert complete · {affected:,} row(s) → `{cloud_schema}.{cloud_table}`."
                    )
                    st.dataframe(cloud_df.head(20), use_container_width=True)
                except Exception as upload_err:
                    st.error(f"Upsert failed: {upload_err}")

        section_label("Database Maintenance")
        st.info(
            "Use these tools to clean your **Neon (PostgreSQL)** database. Fixes data anomalies like duplicates or historical spikes."
        )
        with st.expander("🧼 Scrub / de-duplicate Neon Cloud Database", expanded=False):
            st.markdown(
                "Removes duplicates in Neon tables (PROCESSED_CARD_MONTHLY, etc.) and applies the Yoshinoya normalization fix."
            )
            if st.button(
                "Run cloud scrub / de-duplicate",
                type="primary",
                use_container_width=True,
                disabled=(engine is None),
                key="btn_scrub_neon_cloud",
            ):
                with st.spinner("Cleaning Neon PostgreSQL tables..."):
                    try:
                        from repair_data import scrub_neon_database
                        target_schema = (neon_schema_ingest or "public").strip() or "public"
                        results = scrub_neon_database(engine, schema=target_schema)
                        st.success("✅ Cloud scrub complete!")
                        st.json(results)
                    except Exception as e:
                        st.error(f"Cloud scrub failed: {e}")

        with st.expander("🔴 Danger Zone: Reset Neon Cloud Database", expanded=False):
            st.warning(
                "This will permanently **PURGE ALL DATA** from business, raw, and audit tables in your Neon production database."
            )
            confirm_reset_neon = st.checkbox(
                "I understand this will permanently delete all data in the cloud (PostgreSQL).",
                key="confirm_reset_neon_cloud",
            )
            if st.button(
                "RESET NEON CLOUD DATABASE",
                type="primary",
                disabled=not confirm_reset_neon,
                use_container_width=True,
                key="btn_reset_neon_cloud",
            ):
                with st.spinner("Purging Neon PostgreSQL tables..."):
                    try:
                        from repair_data import reset_neon_database
                        target_schema = (neon_schema_ingest or "public").strip() or "public"
                        results = reset_neon_database(engine, schema=target_schema)
                        st.success("🔥 Neon database reset successfully!")
                        st.json(results)
                    except Exception as e:
                        st.error(f"Reset failed: {e}")

    # ── TAB 2: Audit Log ───────────────────────────────────────────────────────
    with cloud_tab_audit:
        section_label("Recent Ingestion Runs")
        if engine is not None:
            try:
                _aud_schema = st.session_state.get("neon_schema_ingest", "public") or "public"
                st.caption(f"Table `{_aud_schema}.ingestion_runs` — last 10 runs.")
                st.dataframe(
                    fetch_recent_ingestion_runs(engine, schema=_aud_schema, limit=10),
                    use_container_width=True,
                    hide_index=True,
                )
            except Exception as aud_err:
                st.info(f"No ingestion history yet. Run an ingest to populate this log. ({aud_err})")
        else:
            st.warning("Connect to Neon first to view audit logs.")

    st.stop()


def _init_gov_state():
    defaults = {
        "gov_status": "idle",  # idle|blocked|resolved|error
        "gov_delta": {"new_anchors": [], "new_pms": [], "impact_anchor_rows": 0, "impact_pm_rows": 0},
        "gov_decisions": {"approved_anchors": [], "ignored_anchors": [], "approved_pms": [], "ignored_pms": []},
        "gov_signature": None,
        "gov_last_resolved_signature": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _norm_text(value):
    if pd.isna(value):
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned.lower() in {"nan", "none", "null"}:
        return None
    return cleaned


def _read_target_entities(db_path: str) -> pd.DataFrame:
    if not os.path.exists(db_path):
        return pd.DataFrame(columns=["Anchor", "PM"])
    conn = sqlite3.connect(db_path)
    try:
        query = """
            SELECT
                TRIM(MERCHANT_GROUP) AS Anchor,
                TRIM(PM) AS PM
            FROM TARGET
            WHERE MERCHANT_GROUP IS NOT NULL OR PM IS NOT NULL
        """
        return pd.read_sql_query(query, conn)
    except Exception:
        return pd.DataFrame(columns=["Anchor", "PM"])
    finally:
        conn.close()


def _read_master_parameter(path_mon: str) -> pd.DataFrame:
    if not os.path.exists(path_mon):
        return pd.DataFrame(columns=["Anchor", "PM"])
    try:
        wb = openpyxl.load_workbook(path_mon, data_only=True)
        if "PARAMETER" not in wb.sheetnames:
            return pd.DataFrame(columns=["Anchor", "PM"])
        ws = wb["PARAMETER"]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            pm_val = _norm_text(ws.cell(row=row_idx, column=1).value)      # col A
            anchor_val = _norm_text(ws.cell(row=row_idx, column=4).value)  # col D
            if pm_val or anchor_val:
                rows.append({"Anchor": anchor_val, "PM": pm_val})
        return pd.DataFrame(rows, columns=["Anchor", "PM"])
    except Exception:
        return pd.DataFrame(columns=["Anchor", "PM"])


def _detect_governance_delta(db_path: str, path_mon: str) -> dict:
    uploaded_df = _read_target_entities(db_path)
    master_df = _read_master_parameter(path_mon)

    up_anchor_series = uploaded_df["Anchor"].map(_norm_text).dropna() if "Anchor" in uploaded_df.columns else pd.Series(dtype="object")
    up_pm_series = uploaded_df["PM"].map(_norm_text).dropna() if "PM" in uploaded_df.columns else pd.Series(dtype="object")
    m_anchor_series = master_df["Anchor"].map(_norm_text).dropna() if "Anchor" in master_df.columns else pd.Series(dtype="object")
    m_pm_series = master_df["PM"].map(_norm_text).dropna() if "PM" in master_df.columns else pd.Series(dtype="object")

    uploaded_anchors = {x for x in up_anchor_series.tolist() if x}
    uploaded_pms = {x for x in up_pm_series.tolist() if x}
    master_anchors = {x for x in m_anchor_series.tolist() if x}
    master_pms = {x for x in m_pm_series.tolist() if x}

    new_anchors = sorted(uploaded_anchors - master_anchors)
    new_pms = sorted(uploaded_pms - master_pms)

    impact_anchor_rows = int(up_anchor_series.isin(new_anchors).sum()) if len(new_anchors) else 0
    impact_pm_rows = int(up_pm_series.isin(new_pms).sum()) if len(new_pms) else 0
    return {
        "new_anchors": new_anchors,
        "new_pms": new_pms,
        "impact_anchor_rows": impact_anchor_rows,
        "impact_pm_rows": impact_pm_rows,
    }


def _compute_db_signature(db_path: str) -> str:
    if not os.path.exists(db_path):
        return "missing-db"
    stat = os.stat(db_path)
    return f"{int(stat.st_mtime)}-{int(stat.st_size)}"


def _append_to_parameter_sheet(path_mon: str, approved_anchors: list, approved_pms: list):
    if not os.path.exists(path_mon):
        raise FileNotFoundError(f"Master monitoring file not found: {path_mon}")

    wb = openpyxl.load_workbook(path_mon)
    if "PARAMETER" not in wb.sheetnames:
        raise ValueError("Sheet 'PARAMETER' not found in master_monitoring.xlsx")
    ws = wb["PARAMETER"]

    existing_anchor = set()
    existing_pm = set()
    max_data_row = 1

    for row_idx in range(2, ws.max_row + 1):
        pm_val = _norm_text(ws.cell(row=row_idx, column=1).value)
        anchor_val = _norm_text(ws.cell(row=row_idx, column=4).value)
        if pm_val:
            existing_pm.add(pm_val)
            max_data_row = row_idx
        if anchor_val:
            existing_anchor.add(anchor_val)
            max_data_row = row_idx

    # Add anchor rows (A=UNASSIGNED, D=anchor), only net-new
    anchors_to_add = [a for a in approved_anchors if a not in existing_anchor]
    for anchor in anchors_to_add:
        max_data_row += 1
        ws.cell(row=max_data_row, column=1).value = "UNASSIGNED"
        ws.cell(row=max_data_row, column=4).value = anchor
        ws.cell(row=max_data_row, column=2).value = f"=IF(A{max_data_row}=A{max_data_row-1},B{max_data_row-1}+1,1)"
        ws.cell(row=max_data_row, column=3).value = f"=CONCATENATE(A{max_data_row},B{max_data_row})"
        existing_anchor.add(anchor)

    # Add PM rows without forcing anchor assignment, only net-new
    pms_to_add = [p for p in approved_pms if p not in existing_pm]
    for pm in pms_to_add:
        max_data_row += 1
        ws.cell(row=max_data_row, column=1).value = pm
        ws.cell(row=max_data_row, column=4).value = "UNMAPPED_ANCHOR"
        ws.cell(row=max_data_row, column=2).value = f"=IF(A{max_data_row}=A{max_data_row-1},B{max_data_row-1}+1,1)"
        ws.cell(row=max_data_row, column=3).value = f"=CONCATENATE(A{max_data_row},B{max_data_row})"
        existing_pm.add(pm)

    wb.save(path_mon)


def _write_governance_audit(audit_path: str, decisions: dict):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for entity in decisions.get("approved_anchors", []):
        rows.append({"timestamp": now_str, "entity_type": "Anchor", "entity_value": entity, "decision": "approve"})
    for entity in decisions.get("ignored_anchors", []):
        rows.append({"timestamp": now_str, "entity_type": "Anchor", "entity_value": entity, "decision": "ignore"})
    for entity in decisions.get("approved_pms", []):
        rows.append({"timestamp": now_str, "entity_type": "PM", "entity_value": entity, "decision": "approve"})
    for entity in decisions.get("ignored_pms", []):
        rows.append({"timestamp": now_str, "entity_type": "PM", "entity_value": entity, "decision": "ignore"})

    if not rows:
        return

    audit_df = pd.DataFrame(rows)
    if os.path.exists(audit_path):
        old = pd.read_csv(audit_path)
        audit_df = pd.concat([old, audit_df], ignore_index=True)
    Path(audit_path).parent.mkdir(parents=True, exist_ok=True)
    audit_df.to_csv(audit_path, index=False)


page_header(
    "🚀",
    "Automated ETL Pipeline",
    "Local: staging.db + Excel masters + end-to-end analytics (Windows)",
)
st.markdown(
    """<div class="tab-desc">
    Runs on your computer: upload <b>staging.db</b>, set date bounds, resolve governance if needed,
    then run the pipeline that updates <b>master Excel files</b>. (Deployed cloud apps with
    <code>DATABASE_URL</code> use the Neon ingestion page instead.)
    </div>""",
    unsafe_allow_html=True,
)

_init_gov_state()

# ── Pipeline Step Definitions ─────────────────────────────────────────────────
PIPELINE_STEPS = [
    ("📥", "Upload\nDatabase"),
    ("🛠️", "Process\nAnchor MIDs"),
    ("💳", "Card Share\nAnalytics"),
    ("📅", "Weekly\nMonitoring"),
    ("✅", "Complete"),
]

st.markdown("<br>", unsafe_allow_html=True)

@st.dialog("🧪 Quarantine Resolution Required", width="large")
def governance_quarantine_dialog():
    delta = st.session_state.get("gov_delta", {})
    new_anchors = delta.get("new_anchors", [])
    new_pms = delta.get("new_pms", [])

    st.error("Unrecognized Anchors/PMs were detected. Pipeline execution is paused until this is resolved.")
    st.caption(
        f"Impact rows -> Unknown Anchors: {delta.get('impact_anchor_rows', 0)} | "
        f"Unknown PMs: {delta.get('impact_pm_rows', 0)}"
    )

    with st.form("governance_resolution_form"):
        st.markdown("### New Anchors")
        approved_anchors = []
        ignored_anchors = []
        for anchor in new_anchors:
            choice = st.radio(
                f"`{anchor}`",
                options=["Approve & Add to Master", "Ignore/Skip (this run)"],
                horizontal=True,
                key=f"gov_anchor_choice_{anchor}",
            )
            if choice.startswith("Approve"):
                approved_anchors.append(anchor)
            else:
                ignored_anchors.append(anchor)

        st.markdown("### New PMs")
        approved_pms = []
        ignored_pms = []
        for pm in new_pms:
            choice = st.radio(
                f"`{pm}`",
                options=["Approve & Add to Master", "Ignore/Skip (this run)"],
                horizontal=True,
                key=f"gov_pm_choice_{pm}",
            )
            if choice.startswith("Approve"):
                approved_pms.append(pm)
            else:
                ignored_pms.append(pm)

        submit = st.form_submit_button("Submit Resolution", type="primary", use_container_width=True)

    if submit:
        decisions = {
            "approved_anchors": approved_anchors,
            "ignored_anchors": ignored_anchors,
            "approved_pms": approved_pms,
            "ignored_pms": ignored_pms,
        }
        try:
            _append_to_parameter_sheet(PATH_MON, approved_anchors, approved_pms)
            _write_governance_audit(PATH_GOV_AUDIT, decisions)
            st.session_state["gov_decisions"] = decisions
            st.session_state["gov_status"] = "resolved"
            st.session_state["gov_last_resolved_signature"] = st.session_state.get("gov_signature")
            st.success("Governance resolution saved. You can now execute the pipeline.")
            st.rerun()
        except Exception as ex:
            st.session_state["gov_status"] = "error"
            st.error(f"Failed to commit governance decisions: {ex}")

# ── SECTION 1 & 2: Inputs ────────────────────────────────────────────────────
col1, col2 = st.columns([1, 1])

with col1:
    section_label("1. Database Source")
    st.info("Upload your latest `staging.db` containing the raw transaction tables.")

    # ── INGESTION STRATEGY ──
    ingest_strategy = st.radio(
        "Ingestion Strategy",
        ["Replace Full (Slower)", "Update Incremental (Faster)"],
        index=0, horizontal=True,
        help="Full: Overwrites DB. Incremental: Adds only new rows."
    )

    uploaded_db = st.file_uploader("Upload SQLite Database (.db)", type=['db', 'sqlite'])

    if uploaded_db and st.session_state.get("pipeline_step", -1) == -1:
        # 1. ROTATE BACKUPS BEFORE OVERWRITING/UPDATING
        if os.path.exists(PATH_LOCAL_DB):
            rotate_backups(PATH_LOCAL_DB, PATH_BACKUP_DIR)
            
        # 2. SAVE UPLOADED DB
        if ingest_strategy == "Replace Full (Slower)":
            with open(PATH_LOCAL_DB, "wb") as f:
                f.write(uploaded_db.getvalue())
            
            # Update metadata timestamp for full replace
            import sqlite3, datetime
            conn_m = sqlite3.connect(PATH_LOCAL_DB)
            cursor_m = conn_m.cursor()
            cursor_m.execute("CREATE TABLE IF NOT EXISTS APP_METADATA (key TEXT PRIMARY KEY, value TEXT)")
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor_m.execute("INSERT OR REPLACE INTO APP_METADATA (key, value) VALUES ('LAST_DATA_UPDATE', ?)", (now_str,))
            cursor_m.execute("INSERT OR REPLACE INTO APP_METADATA (key, value) VALUES ('NEW_DATA_SIGNAL', '1')")
            conn_m.commit()
            conn_m.close()
            
            st.success(f"✅ `staging.db` replaced! Size: {len(uploaded_db.getvalue()) // (1024*1024)} MB")
        else:
            # Incremental ingestion logic will go here
            temp_path = os.path.join(DB_UPLOAD_DIR, "temp_upload.db")
            with open(temp_path, "wb") as f:
                f.write(uploaded_db.getvalue())
            
            # Placeholder for merge function
            st.info("🔄 Running Incremental Merge (Adding new rows only)...")
            from utils.db_merger import merge_incremental_data
            rows_added = merge_incremental_data(temp_path, PATH_LOCAL_DB)
            os.remove(temp_path)
            st.success(f"✅ Incremental update complete! Added {rows_added} new records.")

        # 3. ADVANCE STEPPER
        st.session_state["pipeline_step"] = 0
        st.session_state["gov_signature"] = _compute_db_signature(PATH_LOCAL_DB)
        st.session_state["gov_delta"] = _detect_governance_delta(PATH_LOCAL_DB, PATH_MON)
        has_delta = bool(st.session_state["gov_delta"]["new_anchors"] or st.session_state["gov_delta"]["new_pms"])
        st.session_state["gov_status"] = "blocked" if has_delta else "resolved"
        st.rerun()
    elif os.path.exists(PATH_LOCAL_DB):
        sz = os.path.getsize(PATH_LOCAL_DB) // (1024 * 1024)
        st.success(f"✅ Existing `staging.db` found locally ({sz} MB).")
    
    else:
        st.warning("⚠️ No database found. Please upload `staging.db`.")

    if os.path.exists(PATH_LOCAL_DB) and st.button("🔍 Re-run Governance Delta Check", use_container_width=True):
        st.session_state["gov_signature"] = _compute_db_signature(PATH_LOCAL_DB)
        st.session_state["gov_delta"] = _detect_governance_delta(PATH_LOCAL_DB, PATH_MON)
        has_delta = bool(st.session_state["gov_delta"]["new_anchors"] or st.session_state["gov_delta"]["new_pms"])
        st.session_state["gov_status"] = "blocked" if has_delta else "resolved"
        st.rerun()
    
    # ── ROLLBACK SECTION ──
    st.markdown("<hr style='margin:1.5rem 0; opacity:0.3;'>", unsafe_allow_html=True)
    st.markdown("### 🔄 Rollback & Restore")
    backups = get_available_backups(PATH_BACKUP_DIR)
    
    if not backups:
        st.caption("No backups available yet.")
    else:
        for b in backups:
            col_b1, col_b2 = st.columns([3, 1])
            col_b1.write(f"**Version {b['version']}** ({b['timestamp']})")
            if col_b2.button(f"Restore", key=f"restore_{b['version']}"):
                if restore_backup(b['path'], PATH_LOCAL_DB):
                    st.success(f"✅ Restored Version {b['version']} successfully!")
                    st.rerun()
                else:
                    st.error("Failed to restore backup.")

    # ── RESET SECTION (New) ──
    st.markdown("<hr style='margin:1.5rem 0; opacity:0.3;'>", unsafe_allow_html=True)
    with st.expander("🗑️ Dangerous Zone: Reset Pipeline"):
        st.warning("This will completely delete the current `staging.db` and all versioned backups.")
        confirm_reset = st.checkbox("I confirm that I want to delete all transaction data.")
        if st.button("🔴 RESET & DELETE ALL DATABASE DATA", type="primary", disabled=not confirm_reset, use_container_width=True):
            try:
                # Delete main DB
                if os.path.exists(PATH_LOCAL_DB):
                    os.remove(PATH_LOCAL_DB)
                
                # Delete backups
                if os.path.exists(PATH_BACKUP_DIR):
                    shutil.rmtree(PATH_BACKUP_DIR)
                    os.makedirs(PATH_BACKUP_DIR)
                
                st.success("🔥 Data wiped successfully! Redirecting...")
                st.session_state["pipeline_step"] = -1
                st.rerun()
            except Exception as e:
                st.error(f"Failed to reset: {e}")

    # ── MAINTENANCE SECTION (New) ──
    st.markdown("<hr style='margin:1.5rem 0; opacity:0.3;'>", unsafe_allow_html=True)
    with st.expander("🛠️ Maintenance & Data Integrity"):
        st.info("Use this if you notice data anomalies or duplicate entries in the dashboard.")
        if st.button("🧼 Scrub/De-duplicate Master Data", use_container_width=True):
            with st.spinner("Cleaning database and Excel files..."):
                try:
                    from repair_data import scrub_database, scrub_excel_card_share, scrub_excel_monitoring
                    
                    scrub_database(PATH_LOCAL_DB)
                    scrub_excel_card_share(PATH_CARD)
                    scrub_excel_monitoring(PATH_MON)
                    
                    st.success("✅ Data scrubbing complete! duplicates removed from Excel and database tables. The Yoshinoya 202503 spike has been normalized.")
                except Exception as e:
                    st.error(f"Scrub failed: {e}")

with col2:
    section_label("2. Extraction Boundaries")

    # Auto-detect the actual date range available in staging.db so the
    # default values always match real data (prevents silent 0-row failures).
    _db_min, _db_max = get_db_date_bounds(PATH_LOCAL_DB)
    today = datetime.today()
    if _db_min and _db_max:
        _default_start = datetime.strptime(_db_min, "%Y-%m-%d").date()
        _default_end   = datetime.strptime(_db_max, "%Y-%m-%d").date()
        st.info(
            f"Auto-detected data window from staging.db: "
            f"**{_db_min}** to **{_db_max}**. Adjust if needed."
        )
    else:
        _default_start = today.replace(day=1).date()
        _default_end   = today.date()
        st.info("Set the SQL date range for the current reporting cycle.")

    start_date = st.date_input("Extract Start Date", value=_default_start)
    end_date   = st.date_input("Extract End Date",   value=_default_end)

st.markdown("<br>", unsafe_allow_html=True)

# ── SECTION 3: Prerequisite Checklist ────────────────────────────────────────
section_label("3. Pre-Flight Check")

prereqs = [
    ("staging.db (Database)",           os.path.exists(PATH_LOCAL_DB)),
    ("master_mid.xlsx (MID Master)",     os.path.exists(PATH_MID)),
    ("master_card_share.xlsx",           os.path.exists(PATH_CARD)),
    ("master_monitoring.xlsx",           os.path.exists(PATH_MON)),
]

all_ready = all(ok for _, ok in prereqs)
gov_blocked = st.session_state.get("gov_status") == "blocked"
gov_delta = st.session_state.get("gov_delta", {})

# Build the checklist HTML
rows_html = "".join([
    f'<div class="prereq-row">'
    f'<span class="prereq-icon">{"✅" if ok else "❌"}</span>'
    f'<span style="{"font-weight:600;" if ok else "opacity:0.7;"}">{label}</span>'
    f'</div>'
    for label, ok in prereqs
])
st.markdown(
    f'<div class="prereq-card">{rows_html}</div>',
    unsafe_allow_html=True,
)

if not all_ready:
    st.error(
        "❌ **Pipeline Locked**: One or more prerequisites are missing. "
        "Upload `staging.db` here and configure Master Excels in **⚙️ Global Settings**."
    )
else:
    st.success("✅ All prerequisites verified. Ready to execute.")

if gov_blocked:
    st.error(
        f"🛑 Governance Gate Active: {len(gov_delta.get('new_anchors', []))} new Anchor(s) "
        f"and {len(gov_delta.get('new_pms', []))} new PM(s) need review."
    )
    if st.button("Open Quarantine Resolution Wizard", type="primary", use_container_width=True):
        governance_quarantine_dialog()
else:
    if st.session_state.get("gov_status") == "resolved":
        st.info("🟢 Governance check passed for current database snapshot.")

# ── SECTION 4: Execute Pipeline ───────────────────────────────────────────────
section_label("4. Execute Automation")

if not is_pipeline_supported():
    st.warning(
        "Cloud runtime detected. Legacy end-to-end pipeline is Windows-only "
        "(uses Excel COM). Use the Cloud Upload & Upsert section above for Neon ingestion."
    )

@st.fragment(run_every="2s")
def execute_pipeline_fragment():
    status_data = get_pipeline_status()
    current_status = status_data.get("status", "idle")
    current_step = status_data.get("step", -1)

    # Always show Stepper at the top of Section 4
    pipeline_stepper(PIPELINE_STEPS, current_step)
    st.markdown("<br>", unsafe_allow_html=True)
    
    if current_status == "idle" or current_status == "":
        if all_ready:
            disable_run = st.session_state.get("gov_status") == "blocked" or (not is_pipeline_supported())
            if disable_run:
                if not is_pipeline_supported():
                    st.warning("Pipeline run is disabled in cloud runtime.")
                else:
                    st.warning("Pipeline is blocked by Governance Gate. Resolve Quarantine first.")
            if st.button(
                "▶️ RUN END-TO-END ANALYTICS PIPELINE",
                type="primary",
                use_container_width=True,
                disabled=disable_run
            ):
                start_str = start_date.strftime("%Y-%m-%d")
                end_str   = end_date.strftime("%Y-%m-%d")
                paths_config = {
                    "PATH_LOCAL_DB": PATH_LOCAL_DB,
                    "PATH_MID": PATH_MID,
                    "PATH_CARD": PATH_CARD,
                    "PATH_MON": PATH_MON,
                    "MASTER_DIR": MASTER_DIR
                }
                start_pipeline_background(start_str, end_str, paths_config)
                st.rerun()

    elif current_status == "running":
        st.info("⏳ **Pipeline is running in background...**")
        
        # Calculate a rough progress percentage based on the step (0-3) vs 4 steps total
        prog = min(current_step / (len(PIPELINE_STEPS) - 1), 1.0) if current_step >= 0 else 0.1
        st.progress(prog)
        
        st.markdown(
            f"**Current Operation:** `{status_data.get('message', 'Processing...')}`\\n\\n"
            "*(This process handles heavy Excel dispatching and hundreds of thousands of classification rules. "
            "Please allow several minutes per step).*"
            "\\n\\nYou can safely navigate to other pages while this runs. "
            "A global notification in the sidebar will alert you when it's finished."
        )
            
    elif current_status == "error":
        st.error(f"❌ **Pipeline Failed:** {status_data.get('error', 'Unknown Error')}")
        if st.button("Acknowledge & Reset"):
            reset_pipeline_status()
            st.rerun()
            
    elif current_status == "complete":
        st.success("🎉 **Pipeline Orchestration Complete!** Your analytics backend is fully synced.")
        step_results = status_data.get("results", {})
        
        p_bg    = "#1A2538"
        p_bdr   = "#2B4470"
        p_txt2  = "#A3B5CC"
        p_gold  = "#F0BE48"
        p_green = "#34D399"

        summary_rows = [
            ("🛠️ Anchor MIDs", f"{step_results.get('mid',{}).get('new', '—')} new classified"),
            ("💳 Card Share records", f"{step_results.get('card',{}).get('rows', '—')} rows"),
            ("📅 Monitoring rows", f"{step_results.get('monitoring',{}).get('rows', '—')} rows"),
            ("🕐 Completed at", status_data.get("start_time", datetime.now().strftime("%d %b %Y %H:%M:%S"))),
        ]
        
        rows_html = "".join([
            f'<div style="display:flex;justify-content:space-between;padding:7px 0;'
            f'border-bottom:1px solid {p_bdr};font-size:0.86rem;">'
            f'<span style="color:{p_txt2};">{k}</span>'
            f'<span style="font-weight:700;color:{p_gold};">{v}</span>'
            f'</div>'
            for k, v in summary_rows
        ])
        
        st.markdown(
            f'''<div style="background:{p_bg};border:1px solid {p_bdr};
                           border-left:4px solid {p_green};border-radius:12px;
                           padding:18px 20px;margin:16px 0;">
              <div style="font-size:0.75rem;text-transform:uppercase;letter-spacing:.07em;
                          color:{p_txt2};margin-bottom:10px;">📋 Pipeline Execution Summary</div>
              {rows_html}
            </div>''',
            unsafe_allow_html=True,
        )

        colA, colB = st.columns(2)
        with colA:
            if st.button("🔄 Reset Pipeline Status", use_container_width=True):
                reset_pipeline_status()
                st.rerun()
        with colB:
            if os.path.exists(PATH_LOCAL_DB):
                st.page_link("pages/4_Dashboard.py", label="**Go to Dashboard 📊**", icon="📈")

execute_pipeline_fragment()
