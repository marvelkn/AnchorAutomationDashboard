import streamlit as st
import sqlite3
import pandas as pd
import openpyxl
import os
import sys

st.set_page_config(page_title="PM Manager", page_icon="👥", layout="wide")

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

from utils.theme import apply_theme, page_header, section_label
apply_theme()

DB_PATH    = os.path.join(_BASE, "database", "staging.db")
EXCEL_PATH = os.path.join(_BASE, "data", "master", "master_monitoring.xlsx")

page_header("👥", "PM Manager", "Manage Project Manager assignments and Merchant Group mappings")

neon_url    = os.getenv("DATABASE_URL")
neon_exists = neon_url is not None

# Guard: if DB is missing, show a clear message instead of crashing.
if not neon_exists and not os.path.exists(DB_PATH):
    st.error(
        "⚠️ **Database not found.** "
        "Please upload `staging.db` via the **Automated Pipeline** page first."
    )
    st.stop()

if "editor_key" not in st.session_state:
    st.session_state.editor_key = 0


# ── Data fetchers ─────────────────────────────────────────────────────────────

def get_cloud_engine():
    from utils.cloud_db import build_engine
    return build_engine()


def fetch_target_data() -> pd.DataFrame:
    if neon_exists:
        engine = get_cloud_engine()
        df = pd.read_sql_query("SELECT merchant_group, pm FROM target ORDER BY pm", engine)
        df.columns = [c.upper() for c in df.columns]
        return df
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT MERCHANT_GROUP, PM FROM TARGET ORDER BY PM", conn)
    conn.close()
    return df


def fetch_all_pms() -> list[str]:
    if neon_exists:
        from sqlalchemy import text
        engine = get_cloud_engine()
        with engine.connect() as conn:
            result = conn.execute(text("SELECT DISTINCT pm FROM target WHERE pm IS NOT NULL ORDER BY pm"))
            return [row[0] for row in result]
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT PM FROM TARGET WHERE PM IS NOT NULL ORDER BY PM")
    pms = [row[0] for row in cursor.fetchall()]
    conn.close()
    return pms


def update_excel_assignment(merchant_group: str, new_pm: str, is_new: bool = False, delete: bool = False):
    """Update master_monitoring.xlsx PARAMETER sheet safely via openpyxl."""
    if not os.path.exists(EXCEL_PATH):
        st.warning(f"Excel file not found at {EXCEL_PATH}. Skipping Excel update.")
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "PARAMETER" not in wb.sheetnames:
            st.warning("PARAMETER sheet not found in master_monitoring.xlsx.")
            return
        sheet = wb["PARAMETER"]
        target_row = None
        max_row = 1
        for row_idx in range(2, sheet.max_row + 1):
            merchant_val = sheet.cell(row=row_idx, column=4).value
            if sheet.cell(row=row_idx, column=1).value is not None:
                max_row = row_idx
            if str(merchant_val).upper().strip() == str(merchant_group).upper().strip():
                target_row = row_idx
                break
        if delete and target_row:
            sheet.cell(row=target_row, column=1).value = "UNASSIGNED"
            wb.save(EXCEL_PATH)
            return
        if target_row:
            sheet.cell(row=target_row, column=1).value = new_pm
        elif is_new:
            new_row_idx = max_row + 1
            sheet.cell(row=new_row_idx, column=1).value = new_pm
            sheet.cell(row=new_row_idx, column=2).value = f"=IF(A{new_row_idx}=A{new_row_idx-1},B{new_row_idx-1}+1,1)"
            sheet.cell(row=new_row_idx, column=3).value = f"=CONCATENATE(A{new_row_idx},B{new_row_idx})"
            sheet.cell(row=new_row_idx, column=4).value = merchant_group
        wb.save(EXCEL_PATH)
    except Exception as e:
        st.error(f"Error updating Excel mapping: {e}")


# ── Load data ─────────────────────────────────────────────────────────────────
try:
    current_data = fetch_target_data()
    all_pms      = fetch_all_pms()
except Exception as e:
    st.error(f"Failed to fetch data: {e}")
    st.stop()

if "UNASSIGNED" not in all_pms:
    all_pms.append("UNASSIGNED")

# ── Summary metrics row ───────────────────────────────────────────────────────
_total_mg    = len(current_data)
_active_pms  = current_data['PM'].nunique() if 'PM' in current_data.columns else 0
_unassigned  = int((current_data['PM'].fillna('UNASSIGNED').str.upper() == 'UNASSIGNED').sum()) if 'PM' in current_data.columns else 0
_avg_per_pm  = round(_total_mg / max(_active_pms, 1), 1)

st.markdown(f"""<div class="stats-grid">
    <div class="stat-card amber">
        <div class="stat-label">Total Merchants</div>
        <div class="stat-value">{_total_mg}</div>
        <div class="stat-meta">merchant groups</div>
    </div>
    <div class="stat-card blue">
        <div class="stat-label">Active PMs</div>
        <div class="stat-value">{_active_pms}</div>
        <div class="stat-meta">project managers</div>
    </div>
    <div class="stat-card {"red" if _unassigned else "green"}">
        <div class="stat-label">Unassigned</div>
        <div class="stat-value">{_unassigned}</div>
        <div class="stat-meta">{"need assignment" if _unassigned else "fully assigned"}</div>
    </div>
    <div class="stat-card purple">
        <div class="stat-label">Avg Merchants / PM</div>
        <div class="stat-value">{_avg_per_pm}</div>
        <div class="stat-meta">per manager</div>
    </div>
</div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── 2-TAB LAYOUT ─────────────────────────────────────────────────────────────
tab_assign, tab_manage = st.tabs(["📋  Assignments", "⚙️  Manage PMs"])


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — ASSIGNMENTS (editable data_editor)
# ─────────────────────────────────────────────────────────────────────────────
with tab_assign:
    st.markdown(
        """<div class="tab-desc">
        Use the <b>Project Manager</b> dropdown in each row to reassign a Merchant Group.
        Changes are only saved when you click <b>Save Assignments</b>.
        </div>""",
        unsafe_allow_html=True,
    )

    edited_df = st.data_editor(
        current_data.copy(),
        column_config={
            "MERCHANT_GROUP": st.column_config.TextColumn(
                "Merchant Group", disabled=True, width="large"
            ),
            "PM": st.column_config.SelectboxColumn(
                "Project Manager", options=all_pms, required=True, width="medium"
            ),
        },
        num_rows="fixed",
        hide_index=True,
        use_container_width=True,
        height=420,
        key=f"data_editor_{st.session_state.editor_key}",
    )

    # Detect changes
    has_changes = not edited_df.equals(current_data)

    if st.button(
        "💾 Save Assignments",
        type="primary",
        disabled=not has_changes,
        use_container_width=False,
        key="btn_save_assignments",
    ):
        diff_mask    = edited_df["PM"] != current_data["PM"]
        changed_rows = edited_df[diff_mask]

        if len(changed_rows) > 0:
            try:
                if neon_exists:
                    from sqlalchemy import text
                    engine = get_cloud_engine()
                    with engine.begin() as conn:
                        for _, row in changed_rows.iterrows():
                            conn.execute(
                                text("UPDATE target SET pm = :pm WHERE merchant_group = :mg"),
                                {"pm": row["PM"], "mg": row["MERCHANT_GROUP"]},
                            )
                            update_excel_assignment(row["MERCHANT_GROUP"], row["PM"])
                else:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    for _, row in changed_rows.iterrows():
                        cursor.execute(
                            "UPDATE TARGET SET PM = ? WHERE MERCHANT_GROUP = ?",
                            (row["PM"], row["MERCHANT_GROUP"]),
                        )
                        update_excel_assignment(row["MERCHANT_GROUP"], row["PM"])
                    conn.commit()
                    conn.close()

                st.success(f"✅ Updated {len(changed_rows)} assignment(s) successfully!")
                st.session_state.editor_key += 1
                st.rerun()
            except Exception as e:
                st.error(f"Error saving assignments: {e}")
    elif not has_changes:
        st.caption("No changes detected. Edit a PM cell in the table above to enable Save.")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — MANAGE PMs (Edit / Add / Danger Zone)
# ─────────────────────────────────────────────────────────────────────────────
with tab_manage:

    # ── Section 1: Edit / Reassign Merchants ─────────────────────────────────
    section_label("✏️ Edit / Reassign Merchants")
    st.markdown(
        """<div class="tab-desc">
        Click any <b>Project Manager</b> cell to reassign a merchant.
        Changes are only saved when you click <b>Save Assignments</b>.
        </div>""",
        unsafe_allow_html=True,
    )

    edited_df_manage = st.data_editor(
        current_data.copy(),
        column_config={
            "MERCHANT_GROUP": st.column_config.TextColumn(
                "Merchant Group", disabled=True, width="large"
            ),
            "PM": st.column_config.SelectboxColumn(
                "Project Manager", options=all_pms, required=True, width="medium"
            ),
        },
        num_rows="fixed",
        hide_index=True,
        use_container_width=True,
        height=420,
        key=f"manage_editor_{st.session_state.editor_key}",
    )

    has_changes_manage = not edited_df_manage.equals(current_data)

    if st.button(
        "💾 Save Assignments",
        type="primary",
        disabled=not has_changes_manage,
        use_container_width=False,
        key="btn_save_manage",
    ):
        diff_mask    = edited_df_manage["PM"] != current_data["PM"]
        changed_rows = edited_df_manage[diff_mask]
        if len(changed_rows) > 0:
            try:
                if neon_exists:
                    from sqlalchemy import text
                    engine = get_cloud_engine()
                    with engine.begin() as conn:
                        for _, row in changed_rows.iterrows():
                            conn.execute(
                                text("UPDATE target SET pm = :pm WHERE merchant_group = :mg"),
                                {"pm": row["PM"], "mg": row["MERCHANT_GROUP"]},
                            )
                            update_excel_assignment(row["MERCHANT_GROUP"], row["PM"])
                else:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    for _, row in changed_rows.iterrows():
                        cursor.execute(
                            "UPDATE TARGET SET PM = ? WHERE MERCHANT_GROUP = ?",
                            (row["PM"], row["MERCHANT_GROUP"]),
                        )
                        update_excel_assignment(row["MERCHANT_GROUP"], row["PM"])
                    conn.commit()
                    conn.close()
                st.success(f"✅ Updated {len(changed_rows)} assignment(s) successfully!")
                st.session_state.editor_key += 1
                st.rerun()
            except Exception as e:
                st.error(f"Error saving assignments: {e}")
    elif not has_changes_manage:
        st.caption("No changes detected. Edit a PM cell in the table above to enable Save.")

    st.divider()

    # ── Section 2: Add New Assignment ─────────────────────────────────────────
    section_label("➕ Add New Assignment")
    with st.form("add_pm_form", clear_on_submit=True):
        new_merchant = st.text_input(
            "Merchant Group Name",
            placeholder="e.g. HYPERMART JAKARTA",
            key="new_mg",
        )
        new_pm_name = st.text_input(
            "Project Manager Name",
            placeholder="e.g. BUDI SANTOSO",
            key="new_pm",
        )
        add_submitted = st.form_submit_button("➕ Add Assignment", type="primary", use_container_width=True)

    if add_submitted:
        new_merchant = new_merchant.strip().upper()
        new_pm_name  = new_pm_name.strip().upper()
        if not new_merchant or not new_pm_name:
            st.warning("Both fields are required.")
        elif new_merchant in current_data["MERCHANT_GROUP"].values:
            st.error("This Merchant Group already exists. Use the Assignments tab to change its PM.")
        else:
            try:
                if neon_exists:
                    from sqlalchemy import text
                    engine = get_cloud_engine()
                    with engine.begin() as conn:
                        conn.execute(
                            text("INSERT INTO target (merchant_group, pm) VALUES (:mg, :pm)"),
                            {"mg": new_merchant, "pm": new_pm_name},
                        )
                else:
                    conn = sqlite3.connect(DB_PATH)
                    conn.execute(
                        "INSERT INTO TARGET (MERCHANT_GROUP, PM) VALUES (?, ?)",
                        (new_merchant, new_pm_name),
                    )
                    conn.commit()
                    conn.close()
                update_excel_assignment(new_merchant, new_pm_name, is_new=True)
                st.success(f"✅ Added **{new_merchant}** → **{new_pm_name}**")
                st.session_state.editor_key += 1
                st.rerun()
            except Exception as e:
                st.error(f"Error adding mapping: {e}")

    st.divider()

    # ── Section 3: Danger Zone ─────────────────────────────────────────────────
    with st.expander("⚠️ Danger Zone: Remove PM", expanded=False):
        _removable = [pm for pm in all_pms if pm != "UNASSIGNED"]
        if not _removable:
            st.info("No PMs available to remove.")
        else:
            pm_to_delete = st.selectbox(
                "PM to Remove",
                options=_removable,
                key="sel_pm_delete",
            )
            reassign_options = [pm for pm in all_pms if pm != pm_to_delete]
            reassign_to = st.selectbox(
                "Reassign their merchants to",
                options=reassign_options,
                key="sel_pm_reassign",
            )
            affected_preview = current_data[current_data["PM"] == pm_to_delete]
            if len(affected_preview) > 0:
                st.warning(
                    f"⚠️ This will reassign **{len(affected_preview)}** merchant(s) "
                    f"currently under **{pm_to_delete}** to **{reassign_to}**."
                )
            else:
                st.info(
                    f"ℹ️ **{pm_to_delete}** has no merchants assigned. "
                    "Removing will only delete them from the PM list."
                )
            if st.button(
                "⚠️ Confirm Removal & Reassign",
                type="primary",
                key="btn_remove_pm",
                use_container_width=True,
            ):
                try:
                    if neon_exists:
                        from sqlalchemy import text
                        engine = get_cloud_engine()
                        with engine.begin() as conn:
                            result = conn.execute(
                                text("SELECT merchant_group FROM target WHERE pm = :pm"),
                                {"pm": pm_to_delete},
                            )
                            affected_merchants = [row[0] for row in result]
                            conn.execute(
                                text("UPDATE target SET pm = :newpm WHERE pm = :oldpm"),
                                {"newpm": reassign_to, "oldpm": pm_to_delete},
                            )
                    else:
                        conn = sqlite3.connect(DB_PATH)
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT MERCHANT_GROUP FROM TARGET WHERE PM = ?", (pm_to_delete,)
                        )
                        affected_merchants = [row[0] for row in cursor.fetchall()]
                        cursor.execute(
                            "UPDATE TARGET SET PM = ? WHERE PM = ?", (reassign_to, pm_to_delete)
                        )
                        conn.commit()
                        conn.close()
                    for merch in affected_merchants:
                        update_excel_assignment(merch, reassign_to)
                    st.success(
                        f"✅ Removed **{pm_to_delete}** and reassigned "
                        f"{len(affected_merchants)} merchant(s) to **{reassign_to}**."
                    )
                    st.session_state.editor_key += 1
                    st.rerun()
                except Exception as e:
                    st.error(f"Error removing PM: {e}")
